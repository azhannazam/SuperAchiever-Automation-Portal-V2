# python_engine/excel_bot.py
import pandas as pd
import os
from datetime import datetime
from database import sync_to_supabase_optimized
import gc
import logging
from pathlib import Path
import re

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)

def parse_ddmmyyyy(date_value):
    """
    Parse dates in DDMMYYYY format (e.g., 22112026 -> 22/11/2026)
    Also handles Excel serial numbers and other formats as fallback
    """
    if pd.isna(date_value) or date_value is None or date_value == '':
        return None
    
    # Convert to string and clean
    date_str = str(date_value).strip()
    
    # Check if it's 8 digits (DDMMYYYY format)
    if len(date_str) == 8 and date_str.isdigit():
        try:
            day = int(date_str[0:2])
            month = int(date_str[2:4])
            year = int(date_str[4:8])
            
            # Validate ranges
            if 1 <= day <= 31 and 1 <= month <= 12 and 2000 <= year <= 2100:
                # Create datetime object
                parsed_date = datetime(year, month, day)
                logger.info(f"    ✅ Parsed DDMMYYYY date: {date_str} -> {parsed_date.strftime('%Y-%m-%d')}")
                return parsed_date.isoformat()
            else:
                logger.warning(f"    ⚠️ Invalid date values: day={day}, month={month}, year={year}")
        except Exception as e:
            logger.warning(f"    ⚠️ Error parsing date '{date_str}': {e}")
    
    # Try Excel serial number format (if it's a number)
    if date_str.replace('.', '').isdigit():
        try:
            from datetime import timedelta
            excel_epoch = datetime(1899, 12, 30)
            days = float(date_str)
            parsed_date = excel_epoch + timedelta(days=days)
            if 2000 <= parsed_date.year <= 2100:
                logger.info(f"    ✅ Parsed Excel serial date: {date_str} -> {parsed_date.strftime('%Y-%m-%d')}")
                return parsed_date.isoformat()
        except:
            pass
    
    # Try standard date parsing as last resort
    try:
        from dateutil import parser
        parsed = parser.parse(date_str, fuzzy=True)
        if 2000 <= parsed.year <= 2100:
            logger.info(f"    ✅ Parsed fuzzy date: {date_str} -> {parsed.strftime('%Y-%m-%d')}")
            return parsed.isoformat()
    except:
        pass
    
    logger.warning(f"    ⚠️ Could not parse date: '{date_str}'")
    return None

def extract_premium_value(row) -> float:
    """
    Extract premium value from row, specifically looking for AFYC column
    """
    # Look specifically for AFYC column
    if 'AFYC' in row and row['AFYC'] is not None:
        value = row['AFYC']
        try:
            if isinstance(value, (int, float)):
                if float(value) > 0:
                    logger.info(f"    ✅ Found premium in AFYC: {value}")
                    return float(value)
            elif isinstance(value, str):
                # Remove any currency symbols and commas
                cleaned = re.sub(r'[RM$,%\s]', '', value)
                if cleaned and cleaned.replace('.', '').isdigit():
                    num_val = float(cleaned)
                    if num_val > 0:
                        logger.info(f"    ✅ Found premium in AFYC: {value} -> {num_val}")
                        return num_val
        except (ValueError, TypeError) as e:
            logger.warning(f"    ⚠️ Could not convert AFYC value '{value}': {e}")
    
    return 0.0

def process_report_316_large_file(file_path, chunk_size=5000):
    """
    Process large Report 316 files using openpyxl
    """
    try:
        from openpyxl import load_workbook
        
        file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
        logger.info(f"🚀 Processing {file_size_mb:.2f}MB file with openpyxl...")
        
        wb = load_workbook(file_path, read_only=True)
        sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]
        logger.info(f"📄 Processing sheet: {sheet_name}")
        
        # Get headers from first row
        headers = []
        for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers.append(cell if cell else f"Column_{len(headers)}")
        
        logger.info(f"📋 Found {len(headers)} columns")
        logger.info(f"📋 Sample headers: {headers[:10]}")
        
        # Find required column indices
        required_columns = {
            'PROPOSALNO': None,
            'PROPOSAL_RECEIVED_DATE': None,  # For submission date
            'RISK_COMMENCEMENT_DATE': None,
            'AGENT_CODE': None,
            'AGENT_NAME': None,
            'AFYC': None,  # For premium
            'PROPOSAL_STATUS': None,
            'AM_NAME': None,
            'POLICYNO': None,
            'PRODUCT_NAME': None,
            'POLY_STATUS': None,
            'PAYMENT_FREQUENCY': None,
            'Factor': None,
            'TOTAL_EXPECTED_DUE': None
        }
        
        for idx, header in enumerate(headers):
            header_upper = str(header).upper().strip()
            
            # Map columns
            if header_upper == 'PROPOSALNO':
                required_columns['PROPOSALNO'] = idx
                logger.info(f"  ✅ Mapped PROPOSALNO to column {idx}")
            elif header_upper == 'PROPOSAL_RECEIVED_DATE':
                required_columns['PROPOSAL_RECEIVED_DATE'] = idx
                logger.info(f"  ✅ Mapped PROPOSAL_RECEIVED_DATE to column {idx}")
            elif header_upper == 'RISK_COMMENCEMENT_DATE':
                required_columns['RISK_COMMENCEMENT_DATE'] = idx
                logger.info(f"  ✅ Mapped RISK_COMMENCEMENT_DATE to column {idx}")
            elif header_upper == 'AGENT_CODE':
                required_columns['AGENT_CODE'] = idx
                logger.info(f"  ✅ Mapped AGENT_CODE to column {idx}")
            elif header_upper == 'AGENT_NAME':
                required_columns['AGENT_NAME'] = idx
                logger.info(f"  ✅ Mapped AGENT_NAME to column {idx}")
            elif header_upper == 'AFYC':
                required_columns['AFYC'] = idx
                logger.info(f"  ✅ Mapped AFYC to column {idx}")
            elif header_upper == 'PROPOSAL_STATUS':
                required_columns['PROPOSAL_STATUS'] = idx
                logger.info(f"  ✅ Mapped PROPOSAL_STATUS to column {idx}")
            elif header_upper == 'AM_NAME':
                required_columns['AM_NAME'] = idx
                logger.info(f"  ✅ Mapped AM_NAME to column {idx}")
            elif header_upper == 'POLICYNO':
                required_columns['POLICYNO'] = idx
                logger.info(f"  ✅ Mapped POLICYNO to column {idx}")
            elif header_upper == 'PRODUCT_NAME':
                required_columns['PRODUCT_NAME'] = idx
                logger.info(f"  ✅ Mapped PRODUCT_NAME to column {idx}")
            elif header_upper == 'POLY_STATUS':
                required_columns['POLY_STATUS'] = idx
                logger.info(f"  ✅ Mapped POLY_STATUS to column {idx}")
            elif header_upper == 'PAYMENT_FREQUENCY':
                required_columns['PAYMENT_FREQUENCY'] = idx
                logger.info(f"  ✅ Mapped PAYMENT_FREQUENCY to column {idx}")
            elif header_upper == 'FACTOR':
                required_columns['Factor'] = idx
                logger.info(f"  ✅ Mapped Factor to column {idx}")
            elif header_upper == 'TOTAL_EXPECTED_DUE':
                required_columns['TOTAL_EXPECTED_DUE'] = idx
                logger.info(f"  ✅ Mapped TOTAL_EXPECTED_DUE to column {idx}")
        
        # Check if essential columns exist
        if required_columns['AGENT_CODE'] is None:
            logger.error("❌ AGENT_CODE column not found")
            wb.close()
            return 0, None
        
        if required_columns['PROPOSALNO'] is None:
            logger.error("❌ PROPOSALNO column not found")
            wb.close()
            return 0, None
        
        logger.info("✅ Required columns found")
        
        # Process in chunks
        chunk_data = []
        total_records = 0
        chunk_number = 0
        matching_records = 0
        
        logger.info("🔍 Scanning for records...")
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row_idx % 10000 == 0:
                logger.info(f"  📊 Scanned {row_idx} rows, found {matching_records} matches...")
            
            if not any(row):
                continue
            
            # Get values using column indices
            proposal_no = row[required_columns['PROPOSALNO']] if required_columns['PROPOSALNO'] is not None else None
            agent_code = row[required_columns['AGENT_CODE']] if required_columns['AGENT_CODE'] is not None else None
            
            if not agent_code or not proposal_no:
                continue
            
            # Get premium value from AFYC column
            afic_value = row[required_columns['AFYC']] if required_columns['AFYC'] is not None else 0
            premium = 0.0
            if afic_value:
                try:
                    if isinstance(afic_value, (int, float)):
                        premium = float(afic_value)
                    elif isinstance(afic_value, str):
                        cleaned = re.sub(r'[RM$,%\s]', '', afic_value)
                        if cleaned and cleaned.replace('.', '').isdigit():
                            premium = float(cleaned)
                except (ValueError, TypeError):
                    premium = 0.0
            
            # Get submission date from PROPOSAL_RECEIVED_DATE and parse it
            submission_date = None
            if required_columns['PROPOSAL_RECEIVED_DATE'] is not None:
                raw_date = row[required_columns['PROPOSAL_RECEIVED_DATE']]
                submission_date = parse_ddmmyyyy(raw_date)
            
            # Get risk commencement date
            risk_date = None
            if required_columns['RISK_COMMENCEMENT_DATE'] is not None:
                raw_risk_date = row[required_columns['RISK_COMMENCEMENT_DATE']]
                risk_date = parse_ddmmyyyy(raw_risk_date)
            
            # Get other values
            agent_name = row[required_columns['AGENT_NAME']] if required_columns['AGENT_NAME'] is not None else 'Unknown'
            proposal_status = row[required_columns['PROPOSAL_STATUS']] if required_columns['PROPOSAL_STATUS'] is not None else ''
            am_name = row[required_columns['AM_NAME']] if required_columns['AM_NAME'] is not None else None
            policy_no = row[required_columns['POLICYNO']] if required_columns['POLICYNO'] is not None else proposal_no
            product_name = row[required_columns['PRODUCT_NAME']] if required_columns['PRODUCT_NAME'] is not None else 'Standard'
            poly_status = row[required_columns['POLY_STATUS']] if required_columns['POLY_STATUS'] is not None else ''
            payment_frequency = row[required_columns['PAYMENT_FREQUENCY']] if required_columns['PAYMENT_FREQUENCY'] is not None else None
            factor = row[required_columns['Factor']] if required_columns['Factor'] is not None else 0
            total_expected_due = row[required_columns['TOTAL_EXPECTED_DUE']] if required_columns['TOTAL_EXPECTED_DUE'] is not None else 0
            
            matching_records += 1
            
            # Convert row to dictionary
            row_dict = {
                'PROPOSALNO': proposal_no,
                'POLICYNO': policy_no,
                'PROPOSAL_RECEIVED_DATE': submission_date,
                'RISK_COMMENCEMENT_DATE': risk_date,
                'AGENT_CODE': agent_code,
                'AGENT_NAME': agent_name,
                'AFYC': premium,
                'PROPOSAL_STATUS': proposal_status,
                'AM_NAME': am_name,
                'PRODUCT_NAME': product_name,
                'POLY_STATUS': poly_status,
                'PAYMENT_FREQUENCY': payment_frequency,
                'Factor': factor,
                'TOTAL_EXPECTED_DUE': total_expected_due
            }
            
            chunk_data.append(row_dict)
            
            if len(chunk_data) >= chunk_size:
                chunk_number += 1
                df_chunk = pd.DataFrame(chunk_data)
                logger.info(f"📦 Processing chunk {chunk_number} with {len(df_chunk)} records")
                
                # Debug: Show premium values in this chunk
                non_zero = len(df_chunk[df_chunk['AFYC'] > 0])
                logger.info(f"  💰 Records with premium > 0: {non_zero}/{len(df_chunk)}")
                
                success, errors = sync_to_supabase_optimized(df_chunk, batch_size=100)
                total_records += success
                
                chunk_data = []
                gc.collect()
        
        if chunk_data:
            chunk_number += 1
            df_chunk = pd.DataFrame(chunk_data)
            logger.info(f"📦 Processing final chunk with {len(df_chunk)} records")
            
            non_zero = len(df_chunk[df_chunk['AFYC'] > 0])
            logger.info(f"  💰 Records with premium > 0: {non_zero}/{len(df_chunk)}")
            
            success, errors = sync_to_supabase_optimized(df_chunk, batch_size=100)
            total_records += success
        
        wb.close()
        
        logger.info(f"📊 Scan complete: {row_idx} total rows scanned")
        logger.info(f"📊 Matching records found: {matching_records}")
        logger.info(f"✅ Successfully synced {total_records} records")
        
        if total_records > 0:
            output_folder = Path("../data/daily_submissions")
            output_folder.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            output_filename = f"Daily_Submissions_Summary_{timestamp}.txt"
            output_path = output_folder / output_filename
            
            with open(output_path, 'w') as f:
                f.write("=" * 50 + "\n")
                f.write("SUPERACHIEVER DAILY SUBMISSION SUMMARY\n")
                f.write("=" * 50 + "\n\n")
                f.write(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Source File: {os.path.basename(file_path)}\n")
                f.write(f"File Size: {file_size_mb:.2f}MB\n\n")
                f.write(f"RESULTS:\n")
                f.write(f"  Total Records Synced: {total_records:,}\n")
                f.write(f"  Matching Records Found: {matching_records:,}\n")
                f.write(f"  Rows Scanned: {row_idx:,}\n")
            
            logger.info(f"📁 Summary file created: {output_filename}")
            return total_records, output_filename
        
        return 0, None
        
    except Exception as e:
        logger.error(f"❌ Error processing large file: {e}")
        import traceback
        traceback.print_exc()
        return 0, None

def process_report_316(file_path):
    """
    Main function to process Report 316 files
    """
    if not os.path.exists(file_path):
        logger.error(f"❌ File not found: {file_path}")
        return 0, None
    
    file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
    logger.info(f"📏 File size: {file_size_mb:.2f}MB")
    
    return process_report_316_large_file(file_path)